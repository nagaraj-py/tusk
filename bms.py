import pandas as pd
import json,csv
from datetime import date
from openpyxl import load_workbook
import re

books_dict={}
ls=[]
ls1=[]
class book_store:
    def __init__(self):
        print("book store")
#take excel file as input and get data from the file to store into data source
    def add_books():
        print("== adding new books ==")
        book = input("enter books file:")
        workbook = load_workbook(filename=book)
        workbook.sheetnames
        sheet = workbook.active

        for value in sheet.iter_rows(values_only=True):
            book=value[0]
            quantity=value[1]
            books_dict[book]=quantity
#storing data from excel file into the datasource file #json file 
        with open('store.json', 'w') as outfile:
                json.dump(books_dict,outfile)

#for add the new book to data source             
    def new_book():
        n = input("enter book :")
        qnt=int(input("enter quantity:"))
        if n not in books_dict.keys():
            print("adding new book ",n)
            books_dict[n] = qnt
        else:
            #books_dict[n] <= qnt
            books_dict[n] =books_dict[n]+qnt

        with open('store.json', 'w') as outfile:
            json.dump(books_dict,outfile)

#store customer details into the cc_data.csv datasource file
    def show_cm_data():
    
        name=input("enter customer name: ").upper()
        try:
            with open('cc_data.csv','r') as csvfile:
        # creating a csv reader object
                csvreader = csv.reader(csvfile)
                for row in csvreader:
                    if re.search(name, str(row)):
                        print(row)
        except:
            print("there is no data on this name")
  
class customer:
#for display the books to use this , read json file show the books in store
    def display_books():
        print("--- Books are in store ---")
        f = open("store.json","r")
        file = f.read()
        obj = json.loads(file)
        books_dict.update(obj)
        df=pd.DataFrame({"menu":books_dict})
        print(df)
#take the inputs from the customers to store into csv data file
    def order_book():
        print("-- customer details --")
        fields=['NAME',' NUMBER ',' DATE ']
        name=input("enter customer name: ").upper()
        num=input("enter customer number: ")       
        print("-- Give book title --")
        choice = input('Enter book title: ')
        if choice in books_dict.keys():
            q=int(input("enter no of copies:"))
            if q <= books_dict[choice]:  
                books_dict[choice] =books_dict[choice]-q
                j=json.dumps(books_dict)
                obj = open('store.json', 'w')
                obj.write(j)
                print("the "+choice," is issued to " +name, "number:"+num)
                #customer data process...
                ls.append("Customer:"+name)
                ls.append(" Number:"+str(num))
                ls.append(" Book:"+str(choice))
                ls.append(" Quantity:"+str(q))
                #ls.append(" Date:"+str(today))
                with open('cc_data.csv', 'a') as csvfile: 
                    csvwriter = csv.writer(csvfile,dialect='excel')
                    csvwriter.writerow(ls)
                    csvfile.close()
                    ls.clear()
            else:
               print("required quantity of books not available in store")
        else:
            print("there is no book on this title")

def main():
    bs=book_store
    cm=customer
    cm.display_books()
    while True:

        print("""
        1.Add books to store
        2.Give book
        3.customer data
        4.exit""")
        opt = int(input("Enter option:"))
        if opt == 1:
            bs.add_books()
        elif opt == 3:
            bs.show_cm_data()
        elif opt == 4:
            break
        elif opt == 2:
            while True:

                print("""== welcome to book store ==
                1.display books
                2.order book
                3.newbook
                4.exit""")
                choice = int(input("enter choice:"))
        
                if choice == 1:
                    cm.display_books()
                elif choice == 2:
                    cm.order_book()
                elif choice==4:
                    break
                elif choice == 3:
                    bs.new_book()
                else:
                    print("Entered wrong option..choose correct option")
        else:
            print("Entered wrong option..choose correct option")
            main()
main()
