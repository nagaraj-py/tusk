import pandas as pd
import json
from openpyxl import load_workbook

books_dict={}

class book_store:
    def __init__(self):
        print("book store")

    def add_books():
        print("== adding new books ==")
        book = input("enter books file:")
        workbook = load_workbook(filename=book)
        workbook.sheetnames
        sheet = workbook.active
        #books_dict={}
        for value in sheet.iter_rows(values_only=True):
            print(value)
            book=value[0]
            quantity=value[1]
            books_dict[book]=quantity
        print(books_dict)

        with open('store.json', 'w') as outfile:
                json.dump(books_dict,outfile)
                
    def new_book():
        n = input("enter book :")
        qnt=int(input("enter quantity:"))
        books_dict[n]=qnt
        with open('store.json', 'w') as outfile:
            json.dump(books_dict,outfile)
        #books.update({n:qnt})

    
class customer:
    def display_books():
        print("--- Books are in store ---")
        f = open("store.json","r")
        file = f.read()
        obj = json.loads(file)
        books_dict.update(obj)
        df=pd.DataFrame({"menu":books_dict})
        print(df)

    def order_book():
        print("-- customer details --")
        name=input("enter customer name: ")
        num=input("enter customer number: ")
        f = open("store.text","a")
        f.write(name)
        f.write(num)
        print("-- Give book title --")
        choice = input('Enter book title: ')
        #q=int(input("enter no of copies:"))
        if choice in books_dict.keys():
            q=input("enter no of copies:")
            #if q in books_dict.items():
            print("the "+choice," is issued to " +name, "number:"+num)
            #init_num = books_dict[choice]
            #init_num -= q
            books_dict[choice] =books_dict[choice]-q
            j=json.dumps(books_dict)
            obj = open('store.json', 'w')
            obj.write(j)
            #else:
             #  print("not available")
        else:
            print("there is no book on this title")

def main():
    bs=book_store
    cm=customer
    while True:

        print("""
        1.Add books to store
        2.Give book
        3.exit""")
        opt = int(input("Enter option:"))
        if opt == 1:
            bs.add_books()
        elif opt == 3:
            break
        elif opt == 2:
            while True:

                print("""== welcome to book store ==
                2.display books
                3.order book
                4.newbook
                5.exit""")
                choice = int(input("enter choice:"))
        
                if choice == 2:
                    cm.display_books()
                elif choice == 3:
                    cm.order_book()
                elif choice==5:
                    break
                elif choice == 4:
                    bs.new_book()
                else:
                    continue
main()